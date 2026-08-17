#!/usr/bin/env python
"""企业高管背调：负面内容台账 生成/追加 脚本（Excel）。

用法：
  python build_ledger.py <entries.json> <output.xlsx>              # 新建台账
  python build_ledger.py <entries.json> <existing.xlsx> --append   # 追加到已有台账并重排

entries.json 格式（数组，每项一个台账行；未提供的字段留空）：
[
  {
    "分组": "A 建议处置",            # A 建议处置 / B 不建议投诉
    "平台/账号": "微博·@xxx",
    "标题/内容": "……",
    "链接": "https://……",
    "发布/时间": "2026-02-03 23:09",
    "内容定性": "L1 攻击性",          # L1 攻击性 / L2 失实传闻 / L3 客观事实
    "主要侵权/违规点": "……",
    "建议处置方式": "……",
    "处置优先级": "P1 优先",          # P0 立即 / P1 优先 / P2 常规 / P3 观察 / —
    "预期成功率": "中",               # 高 / 中 / 中低 / 低 / —
    "处置状态": "待处置",
    "责任人": "",
    "处置日期": "",
    "备注": "来源：……"
  }
]

样式规范（沿用既有台账样式）：
- 表头：深蓝底(1F3864)白字加粗，行高 26
- 分组列(B)：A=浅绿(E8F5E9)，B=浅灰蓝(EEF2F8)
- 优先级列(J)：P0=浅红(FDE9E8) P1=浅橙(FDF0E4) P2=浅黄(FDF7E6) P3=灰蓝(EEF2F8)
- 数据行：微软雅黑 10、thin 全边框、行高 70、冻结首行、自动筛选
- 排序：分组 → 处置优先级 → 预期成功率，序号重排
"""
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side

HEADERS = ["序号", "分组", "平台/账号", "标题/内容", "链接", "发布/时间", "内容定性",
           "主要侵权/违规点", "建议处置方式", "处置优先级", "预期成功率", "处置状态",
           "责任人", "处置日期", "备注"]

GRP_ORDER = {"A 建议处置": 0, "B 不建议投诉": 1}
PRI_ORDER = {"P0 立即": 0, "P1 优先": 1, "P2 常规": 2, "P3 观察": 3, "—": 4}
SUC_ORDER = {"高": 0, "中": 1, "中低": 2, "低": 3, "—": 4}

FILL_HEADER = PatternFill("solid", fgColor="1F3864")
FILL_GROUP_A = PatternFill("solid", fgColor="E8F5E9")
FILL_GROUP_B = PatternFill("solid", fgColor="EEF2F8")
FILL_PRI = {0: PatternFill("solid", fgColor="FDE9E8"),
            1: PatternFill("solid", fgColor="FDF0E4"),
            2: PatternFill("solid", fgColor="FDF7E6"),
            3: PatternFill("solid", fgColor="EEF2F8")}
FILL_NONE = PatternFill(fill_type=None)
THIN = Border(*(Side(style="thin", color="B0B0B0"),) * 4)
WIDTHS = {"A": 6, "B": 13, "C": 22, "D": 36, "E": 44, "F": 12, "G": 11,
          "H": 42, "I": 28, "J": 11, "K": 10, "N": 12, "O": 36}


def load_entries(path):
    # utf-8-sig 容忍带 BOM 的输入（PowerShell Out-File 等工具可能写入 BOM）
    with open(path, encoding="utf-8-sig") as f:
        data = json.load(f)
    rows = []
    for e in data:
        rows.append([e.get(h, "") for h in HEADERS[1:]])  # 序号由脚本统一重排
    return rows


def style_header(ws):
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFFFF")
        cell.fill = FILL_HEADER
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN
    ws.row_dimensions[1].height = 26


def write_rows(ws, rows):
    # rows 为 14 字段（HEADERS[1:]）：r[0]=分组 r[1]=平台 r[2]=标题 … r[8]=处置优先级 r[9]=预期成功率
    rows.sort(key=lambda r: (GRP_ORDER.get(r[0], 9), PRI_ORDER.get(r[8], 4), SUC_ORDER.get(r[9], 4)))
    for i, row in enumerate(rows, 1):
        cells = [i] + row  # 序号 + 14 字段 = 15 列
        r = 1 + i
        ws.row_dimensions[r].height = 70
        for c, v in enumerate(cells, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(name="微软雅黑", size=10)
            cell.border = THIN
            cell.alignment = openpyxl.styles.Alignment(vertical="top", wrap_text=c in (4, 8, 15))
            if c == 2:
                cell.fill = FILL_GROUP_A if cells[1] == "A 建议处置" else FILL_GROUP_B
            elif c == 10:
                fill = FILL_PRI.get(PRI_ORDER.get(cells[9], 4))
                cell.fill = fill if fill else FILL_NONE
            else:
                cell.fill = FILL_NONE


def finalize(ws, n_rows):
    for col, w in WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:O{1 + n_rows}"


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(2)
    entries_path, out_path = sys.argv[1], sys.argv[2]
    append = "--append" in sys.argv[3:]

    new_rows = load_entries(entries_path)

    if append and Path(out_path).exists():
        wb = openpyxl.load_workbook(out_path)
        ws = wb["处置台账"]
        rows = []
        for r in range(2, ws.max_row + 1):
            vals = [ws.cell(row=r, column=c).value for c in range(2, 16)]
            if all(v in (None, "") for v in vals):
                continue
            rows.append(vals)
        rows.extend(new_rows)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "处置台账"
        style_header(ws)
        rows = list(new_rows)

    write_rows(ws, rows)
    finalize(ws, len(rows))
    wb.save(out_path)
    print(f"已保存 {len(rows)} 条 → {out_path}")


if __name__ == "__main__":
    main()
